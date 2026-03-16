from flask import Flask, request, jsonify, send_file, render_template, session, redirect
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import io, os, sqlite3, hashlib, secrets
from collections import defaultdict
from functools import wraps

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))
DB_PATH = os.environ.get('DB_PATH', 'gst_portal.db')

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def init_db():
    conn = get_db(); c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL, password_hash TEXT NOT NULL,
        full_name TEXT, role TEXT DEFAULT 'user',
        created_at TEXT DEFAULT CURRENT_TIMESTAMP)''')
    c.execute('''CREATE TABLE IF NOT EXISTS hotels (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL, code TEXT UNIQUE NOT NULL, gstin TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP)''')
    c.execute('''CREATE TABLE IF NOT EXISTS saved_files (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hotel_id INTEGER NOT NULL, file_type TEXT NOT NULL,
        filename TEXT NOT NULL, file_data BLOB NOT NULL, file_size INTEGER,
        uploaded_by INTEGER, uploaded_at TEXT DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(hotel_id, file_type))''')
    c.execute('''CREATE TABLE IF NOT EXISTS history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL, module TEXT NOT NULL,
        step TEXT NOT NULL, filename TEXT NOT NULL,
        processed_at TEXT DEFAULT CURRENT_TIMESTAMP)''')
    if not c.execute("SELECT id FROM users WHERE username='admin'").fetchone():
        c.execute("INSERT INTO users (username,password_hash,full_name,role) VALUES (?,?,?,?)",
                  ('admin', hash_pw('admin123'), 'Administrator', 'admin'))
        c.execute("INSERT OR IGNORE INTO hotels (name,code,gstin) VALUES (?,?,?)",
                  ('TWG Hotels - Jaipur Marriott', 'JMH', ''))
    conn.commit(); conn.close()

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            if request.method == 'POST':
                return jsonify({'error':'Login required'}), 401
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated

def get_hotel_id():
    hid = int(request.form.get('hotel_id', 0))
    if hid == 0:
        conn = get_db()
        h = conn.execute("SELECT id FROM hotels LIMIT 1").fetchone()
        conn.close()
        hid = h['id'] if h else 1
    return hid

def get_file_bytes(hotel_id, file_type, req_key):
    f = request.files.get(req_key)
    if f and f.filename: return f.read()
    conn = get_db()
    row = conn.execute("SELECT file_data FROM saved_files WHERE hotel_id=? AND file_type=?",
                       (hotel_id, file_type)).fetchone()
    conn.close()
    return bytes(row['file_data']) if row else None

def save_history(module, step, filename):
    try:
        conn = get_db()
        conn.execute("INSERT INTO history (user_id,module,step,filename) VALUES (?,?,?,?)",
                     (session.get('user_id',1), module, step, filename))
        conn.commit(); conn.close()
    except: pass

# ── Exact formats from original ───────────────────────────
FMT_COMMA = '_ * #,##0.00_ ;_ * \\-#,##0.00_ ;_ * \\-??_ ;_ @_ '
FMT_INT   = '#,##0'
FMT_PCT   = '0%'

def hdr_font():  return Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')
def hdr_fill():  return PatternFill('solid', fgColor='FF000000')
def data_font(): return Font(name='Calibri', size=10)
def bold_font(): return Font(name='Calibri', size=10, bold=True)

def apply_header(ws, row, start_col, end_col):
    for c in range(start_col, end_col + 1):
        ws.cell(row, c).font = hdr_font()
        ws.cell(row, c).fill = hdr_fill()

def apply_total_row(ws, row, start_col, end_col):
    for c in range(start_col, end_col + 1):
        ws.cell(row, c).font = bold_font()
        ws.cell(row, c).fill = hdr_fill()

def set_heights(ws, last_row):
    for r in range(1, last_row + 2):
        ws.row_dimensions[r].height = 13

# ══════════════════════════════════════════════════════════
# READ EXCEL
# ══════════════════════════════════════════════════════════
def read_wb(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    out = {}
    for name in wb.sheetnames:
        out[name] = [list(r) for r in wb[name].iter_rows(values_only=True)]
    return out

def get_sheet(data, *names):
    for n in names:
        if n in data:
            return data[n]
    return list(data.values())[0]

# ══════════════════════════════════════════════════════════
# BUILD MAPPING from Backend sheet
# Row 1 = date, Row 2 = headers, Data from Row 3
# Cols: B=TRX_DESC, C=Tax Head, D=Rate, E=HSN Grouping, F=HSN Code, G=BOF Code
# ══════════════════════════════════════════════════════════
def build_map(backend_rows):
    m = {}
    for row in backend_rows[2:]:
        if not row or not row[1]:
            continue
        key = str(row[1]).strip().lower()
        m[key] = {
            'taxHead':  str(row[2]).strip()  if len(row) > 2 and row[2] else '',
            'rate':     row[3]               if len(row) > 3 and row[3] is not None else '',
            'hsnDesc':  str(row[4]).strip()  if len(row) > 4 and row[4] else '',
            'hsnCode':  str(row[5]).strip()  if len(row) > 5 and row[5] else '',
            'bofCode':  str(row[6]).strip()  if len(row) > 6 and row[6] else '',
        }
    return m

def lookup(desc, m):
    k = str(desc).strip().lower()
    if k in m:
        return m[k]
    for mk, mv in m.items():
        if k and mk and (k in mk or mk in k):
            return mv
    return None

# ══════════════════════════════════════════════════════════
# BUILD GSTR-4A EINV MAP
# bill_no → {gstin, name, inv_type, taxable}
# ══════════════════════════════════════════════════════════
def build_einv_map(gstr_rows):
    m = {}
    if not gstr_rows or len(gstr_rows) < 2:
        return m
    hdrs = gstr_rows[0]
    ci = {str(h).strip(): i for i, h in enumerate(hdrs) if h}
    for row in gstr_rows[1:]:
        if not row:
            continue
        try:
            bill = str(row[ci.get('Invoice number', 2)] or '').strip()
            if not bill:
                continue
            if bill not in m:
                m[bill] = {
                    'gstin':    str(row[ci.get('GSTIN/UIN of Recipient', 0)] or '').strip(),
                    'name':     str(row[ci.get('Receiver Name', 1)] or '').strip(),
                    'inv_type': str(row[ci.get('Invoice Type', 8)] or 'Regular B2B').strip(),
                    'taxable':  0,
                }
            m[bill]['taxable'] += float(row[ci.get('Taxable Value', 11)] or 0)
        except:
            pass
    return m

# ══════════════════════════════════════════════════════════
# PROCESS D110
# Adds cols O-Y exactly as VBA:
# O=Tax Head, P=Rate, Q=HSN Description, R=HSN Code
# S=Rate Check, T=Invoice Check
# U=GSTIN, V=Receiver Name, W=Invoice Type, X=Document Type, Y=Supply Type
# ══════════════════════════════════════════════════════════
def process_d110(d110_rows, mapping, einv_map):
    if not d110_rows:
        return [], []

    # Trim trailing None/blank cols — D110 Dr.Base has 5 empty cols after N
    headers = list(d110_rows[0])
    while headers and headers[-1] is None:
        headers.pop()
    # Keep only first 14 cols (A-N: FOLIO_TYPE to TRANSACTION_DESCRIPTION)
    NCOLS = 14
    headers = headers[:NCOLS]

    ci = {str(h).strip(): i for i, h in enumerate(headers) if h}

    new_hdrs = headers + [
        'Tax Head', 'Rate', 'HSN Description', 'HSN Code',
        'Rate Check', 'Invoice Check',
        'GSTIN', 'Receiver Name', 'Invoice Type', 'Document Type', 'Supply Type'
    ]

    # First pass — map each row + accumulate bill totals for Rate Check
    bill_rate_check = defaultdict(float)
    bill_debit_sum  = defaultdict(float)
    temp = []

    for row in d110_rows[1:]:
        if not row or all(c is None for c in row):
            continue

        # Trim row to 14 cols (same as headers — ignore extra blank cols from source)
        row = list(row[:NCOLS])

        desc = str(row[ci.get('TRANSACTION_DESCRIPTION', 13)] or '').strip()
        mp   = lookup(desc, mapping)

        bill = str(row[ci.get('BILL_NO', 1)] or '').strip()

        try: debit = float(row[ci.get('FT_DEBIT', 11)] or 0)
        except: debit = 0

        tax_head = mp['taxHead'] if mp else 'UNMAPPED'
        rate     = mp['rate']    if mp else ''
        hsn_desc = mp['hsnDesc'] if mp else ''
        hsn_code = mp['hsnCode'] if mp else ''

        # Rate Check (col S) — VBA: =IFS(RC[-3]=0,RC[-7]*-1,RC[-3]="No GST",0,RC[-3]<>0,RC[-7]/RC[-3])
        # NON GST / OTHERS (rate="NO GST") -> 0
        # REVENUE (rate blank) -> debit * -1
        # Tax rows -> debit / rate
        try:
            rate_str = str(rate).strip().upper() if rate != '' else ''
            if rate_str == 'NO GST':
                rate_check = 0
            elif rate_str == '' or rate_str == '0':
                rate_check = debit * -1
            else:
                rate_num = float(rate)
                rate_check = debit / rate_num if rate_num else debit * -1
        except:
            rate_check = 0

        bill_rate_check[bill] += rate_check
        bill_debit_sum[bill]  += debit

        # EINV lookup (cols U-Y)
        einv = einv_map.get(bill, {})
        gstin    = einv.get('gstin', 'B2C')
        rec_name = einv.get('name',  'B2C')
        inv_type = einv.get('inv_type', 'Regular B2C')

        temp.append((bill, list(row) + [
            tax_head, rate, hsn_desc, hsn_code,
            rate_check, None,           # Invoice Check = pass 2
            gstin, rec_name, inv_type, None, None  # DocType, SupplyType = pass 2
        ]))

    # Second pass — fill Invoice Check, Document Type, Supply Type
    O = len(headers)  # offset for new cols
    processed = []
    for bill, row in temp:
        rc_sum   = bill_rate_check.get(bill, 0)
        inv_check = 'OK' if abs(rc_sum) <= 50 else 'Error'   # col T (O+5)
        doc_type  = 'Credit Note' if bill_debit_sum.get(bill, 0) < 0 else 'Invoice'  # col X (O+9)
        gstin_val = row[O + 6]
        supply    = 'B2C' if not gstin_val or gstin_val == 'B2C' else 'B2B'          # col Y (O+10)

        row[O + 5]  = inv_check
        row[O + 9]  = doc_type
        row[O + 10] = supply
        processed.append(row)

    return processed, new_hdrs

# ══════════════════════════════════════════════════════════
# STEP 1 — PROC D110
# ══════════════════════════════════════════════════════════
def make_proc_d110(d110_rows, mapping, einv_map):
    processed, new_hdrs = process_d110(d110_rows, mapping, einv_map)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Proc_D110'

    for c, h in enumerate(new_hdrs, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(name='Calibri', size=10, bold=True)

    for r, row in enumerate(processed, 2):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v).font = data_font()

    set_heights(ws, len(processed) + 1)
    ws.sheet_view.showGridLines = False
    return wb, processed, new_hdrs

# ══════════════════════════════════════════════════════════
# STEP 2 — SALES REGISTER
# Pivot rows: GSTIN | Rec Name | Bill_no | Date | HSN Desc | HSN Code |
#             Invoice Type | Doc Type | Supply Type | Rate
# Pivot cols: IGST | CGST | SGST | CESS  (REVENUE/NON GST hidden)
# Extra cols: Revenue as per Taxes | Invoice Revenue | Revenue as per D110 |
#             Difference | Status
# ══════════════════════════════════════════════════════════
def make_sales_register(processed, new_hdrs, einv_map):
    ci = {str(h).strip(): i for i, h in enumerate(new_hdrs) if h}

    TAX_HEADS = ['IGST', 'CGST', 'SGST', 'CESS']

    # Pivot accumulation
    pivot       = defaultdict(lambda: defaultdict(float))
    bill_rev     = defaultdict(float)  # Revenue per bill_no (for Rev as per D110)

    for row in processed:
        th = str(row[ci.get('Tax Head', 0)] or '').strip().upper()

        gstin       = str(row[ci.get('GSTIN', 0)]            or '').strip()
        rec_name    = str(row[ci.get('Receiver Name', 0)]    or '').strip()
        bill_no     = str(row[ci.get('BILL_NO', 1)]          or '').strip()
        bill_date   = row[ci.get('BILL_GENERATION_DATE', 2)]
        hsn_desc    = str(row[ci.get('HSN Description', 0)]  or '').strip()
        hsn_code    = str(row[ci.get('HSN Code', 0)]         or '').strip()
        inv_type    = str(row[ci.get('Invoice Type', 0)]     or 'Regular B2C').strip()
        doc_type    = str(row[ci.get('Document Type', 0)]    or 'Invoice').strip()
        supply_type = str(row[ci.get('Supply Type', 0)]      or 'B2C').strip()
        rate        = row[ci.get('Rate', 0)]

        try: debit = float(row[ci.get('FT_DEBIT', 11)] or 0)
        except: debit = 0

        key = (gstin, rec_name, bill_no, bill_date, hsn_desc, hsn_code,
               inv_type, doc_type, supply_type, rate)

        if th in TAX_HEADS:
            pivot[key][th] += debit
        elif th == 'REVENUE':
            bill_rev[bill_no] += debit

    # Write Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sales_Register'

    # Col widths from original
    for col, w in [('A',19.42),('B',55.29),('C',9.86),('D',22.57),('E',16.85),
                   ('F',11.0),('G',24.86),('H',15.71),('I',12.57),('J',10.0),
                   ('K',10.29),('L',13.71),('M',10.0),('N',7.29),
                   ('O',18.0),('P',14.71),('Q',17.86),('R',12.71),('S',7.29)]:
        ws.column_dimensions[col].width = w

    # Row 3 = GST Amount label (matches original pivot layout)
    ws.cell(3, 11, 'GST Amount').font = data_font()

    # Row 4 = column headers (black/white)
    pivot_hdrs = [
        'GSTIN', 'Receiver Name', 'Bill_no', 'BILL_GENERATION_DATE',
        'HSN Description', 'HSN Code', 'Invoice Type', 'Document Type',
        'Supply Type', 'Rate',
        'IGST', 'CGST', 'SGST', 'CESS',
        'Revenue as per Taxes', 'Invoice Revenue',
        'Revenue as per D110', 'Difference', 'Status'
    ]
    for c, h in enumerate(pivot_hdrs, 1):
        ws.cell(4, c, h)
    apply_header(ws, 4, 1, len(pivot_hdrs))

    DATA_START = 5
    rows_written = 0

    for key, tax_vals in pivot.items():
        gstin, rec_name, bill_no, bill_date, hsn_desc, hsn_code, \
            inv_type, doc_type, supply_type, rate = key

        igst = tax_vals.get('IGST', 0) or 0
        cgst = tax_vals.get('CGST', 0) or 0
        sgst = tax_vals.get('SGST', 0) or 0
        cess = tax_vals.get('CESS', 0) or 0

        # Col O: Revenue as per Taxes = SUM(IGST:CESS) / Rate  [VBA: =SUM(K:N)/J]
        try:
            rate_num = float(rate) if rate else 0
            rev_per_tax = (igst + cgst + sgst + cess) / rate_num if rate_num else 0
        except:
            rev_per_tax = 0

        # Col Q: Revenue as per D110 = SUMIFS(Proc_D110!L, O="revenue", B=bill_no)
        rev_d110 = bill_rev.get(bill_no, 0)

        # Col P: Invoice Revenue = SUMIF(C:C, bill_no, O:O)
        # = Sum of "Revenue as per Taxes" for ALL rows with same bill_no
        # This gets filled AFTER all rows written (second pass below)
        inv_rev = 0  # placeholder - filled in second pass

        # Col R: Difference = P - Q  (Invoice Revenue - Revenue as per D110)
        # Col S: Status
        diff   = 0   # placeholder
        status = ''  # placeholder'

        r = DATA_START + rows_written
        row_data = [
            gstin, rec_name, bill_no, bill_date,
            hsn_desc, hsn_code, inv_type, doc_type, supply_type, rate,
            igst if igst else None,
            cgst if cgst else None,
            sgst if sgst else None,
            cess if cess else None,
            rev_per_tax, inv_rev, rev_d110, diff, status
        ]

        for c, v in enumerate(row_data, 1):
            cell = ws.cell(r, c, v)
            cell.font = data_font()
            if c in [11, 12, 13, 14, 15, 16, 17, 18]:
                cell.number_format = FMT_COMMA
            # col P(16), R(18), S(19) filled correctly in second pass below

        rows_written += 1

    LAST_DATA = DATA_START + rows_written - 1

    # ── Second pass: Fill col P (Invoice Revenue) = SUMIF(C:C, bill_no, O:O)
    # Build bill_no → sum of rev_per_tax (col O) across all rows for that bill
    bill_inv_rev = {}  # bill_no → sum of col O
    for row_num in range(DATA_START, LAST_DATA + 1):
        bill_val = ws.cell(row_num, 3).value   # col C = Bill_no
        o_val    = ws.cell(row_num, 15).value  # col O = Revenue as per Taxes
        if bill_val:
            key = str(bill_val)
            try:    bill_inv_rev[key] = bill_inv_rev.get(key, 0) + float(o_val or 0)
            except: pass

    # Write col P, R, S for every data row
    for row_num in range(DATA_START, LAST_DATA + 1):
        bill_val = str(ws.cell(row_num, 3).value or '')
        q_val    = ws.cell(row_num, 17).value  # col Q = Revenue as per D110

        p_val = bill_inv_rev.get(bill_val, 0)  # col P = Invoice Revenue
        ws.cell(row_num, 16, p_val).number_format = FMT_COMMA

        try:    diff = p_val - float(q_val or 0)
        except: diff = 0
        ws.cell(row_num, 18, diff).number_format = FMT_COMMA  # col R

        status = 'Passed' if abs(diff) <= 50 else 'Pending'
        s_cell = ws.cell(row_num, 19, status)  # col S
        if status == 'Passed':
            s_cell.font = Font(name='Calibri', size=10, color='0070C0')
        else:
            s_cell.font = Font(name='Calibri', size=10, color='C00000')

    TOTAL_ROW = LAST_DATA + 1
    ws.cell(TOTAL_ROW, 1, 'TOTAL')
    for c in range(7, len(pivot_hdrs) + 1):
        cl = get_column_letter(c)
        cell = ws.cell(TOTAL_ROW, c, f'=SUM({cl}{DATA_START}:{cl}{LAST_DATA})')
        cell.number_format = FMT_COMMA
    apply_total_row(ws, TOTAL_ROW, 1, len(pivot_hdrs))

    set_heights(ws, TOTAL_ROW)
    ws.sheet_view.showGridLines = False
    return wb

# ══════════════════════════════════════════════════════════
# STEP 3 — SR NO GST
# Pivot: Bill_no | Rate | HSN Description × SUM FT_DEBIT
# Filter: Rate = "NO GST" ONLY
# ══════════════════════════════════════════════════════════
def make_sr_no_gst(processed, new_hdrs):
    ci = {str(h).strip(): i for i, h in enumerate(new_hdrs) if h}

    pivot = defaultdict(float)

    for row in processed:
        rate = str(row[ci.get('Rate', 0)] or '').strip().upper()
        # VBA: only show Rate = "NO GST"
        if rate != 'NO GST':
            continue

        bill_no = str(row[ci.get('BILL_NO', 1)] or '').strip()

        try: debit = float(row[ci.get('FT_DEBIT', 11)] or 0)
        except: debit = 0

        # VBA groups by Bill_no (Rate always NO GST, HSN Desc subtotals OFF)
        pivot[bill_no] += debit

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'SR NO GST'

    for col, w in [('A',10.57),('B',9.57),('C',17.71),('D',12.0),('E',9.14)]:
        ws.column_dimensions[col].width = w

    # Rows 3-4 = headers (VBA pivot creates 2 header rows), data row 5
    for c, h in enumerate(['Bill_no','Rate','HSN Description','DR Amount'], 1):
        ws.cell(3, c, h)
    apply_header(ws, 3, 1, 4)

    for c, h in enumerate(['Bill_no','Rate','HSN Description','DR Amount'], 1):
        ws.cell(4, c, h)
    apply_header(ws, 4, 1, 4)

    DATA_START = 5
    r = DATA_START

    for bill_no, amt in sorted(pivot.items()):
        ws.cell(r, 1, bill_no).font  = data_font()
        ws.cell(r, 2, 'NO GST').font = data_font()
        ws.cell(r, 3, 0).font        = data_font()
        cell = ws.cell(r, 4, amt)
        cell.font = data_font()
        cell.number_format = FMT_INT
        r += 1

    LAST_DATA = r - 1
    TOTAL_ROW = r

    ws.cell(TOTAL_ROW, 1, 'TOTAL')
    total_amt = sum(v for v in pivot.values())
    cell = ws.cell(TOTAL_ROW, 4, total_amt)
    cell.number_format = FMT_INT
    apply_total_row(ws, TOTAL_ROW, 1, 4)

    set_heights(ws, TOTAL_ROW)
    ws.sheet_view.showGridLines = False
    return wb

# ══════════════════════════════════════════════════════════
# STEP 4 — HSN SUMMARY
# Pivot: HSN Code | HSN Description | Rate × Tax Head cols (IGST/CGST/SGST/CESS)
# Extra col H = Revenue = IFERROR(SUM(IGST:CESS)/Rate,"")
# ══════════════════════════════════════════════════════════
def make_hsn_summary(processed, new_hdrs):
    ci = {str(h).strip(): i for i, h in enumerate(new_hdrs) if h}

    TAX_HEADS = ['IGST', 'CGST', 'SGST', 'CESS']

    pivot = defaultdict(lambda: defaultdict(float))

    for row in processed:
        th = str(row[ci.get('Tax Head', 0)] or '').strip().upper()
        if th not in TAX_HEADS:
            continue

        hsn_code = str(row[ci.get('HSN Code', 0)] or '').strip()
        hsn_desc = str(row[ci.get('HSN Description', 0)] or '').strip()
        rate     = row[ci.get('Rate', 0)]

        try: debit = float(row[ci.get('FT_DEBIT', 11)] or 0)
        except: debit = 0

        pivot[(hsn_code, hsn_desc, rate)][th] += debit

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'HSN Summary'

    for col, w in [('A',11.0),('B',17.29),('C',6.85),('D',11.57),
                   ('E',10.29),('F',10.29),('G',4.57),('H',11.29)]:
        ws.column_dimensions[col].width = w

    # Row 3 = header row (black/white — full)
    for c, h in enumerate(['HSN Code','HSN Description','Rate','IGST','CGST','SGST','CESS','Revenue'], 1):
        ws.cell(3, c, h)
    apply_header(ws, 3, 1, 8)

    # Row 4 = sub-header for pivot tax cols only
    for c, h in enumerate(['HSN Code','HSN Description','Rate','IGST','CGST','SGST','CESS','Revenue'], 1):
        ws.cell(4, c, h)
    apply_header(ws, 4, 4, 8)

    DATA_START = 5
    r = DATA_START

    for (hsn_code, hsn_desc, rate), tax_vals in sorted(pivot.items(), key=lambda x: (str(x[0][0]), str(x[0][2]))):
        igst = tax_vals.get('IGST', 0) or 0
        cgst = tax_vals.get('CGST', 0) or 0
        sgst = tax_vals.get('SGST', 0) or 0
        cess = tax_vals.get('CESS', 0) or 0

        # Skip rows where everything is zero AND HSN code is blank
        if not hsn_code and igst == 0 and cgst == 0 and sgst == 0 and cess == 0:
            continue
        # Revenue back-calc: IFERROR(SUM(IGST:CESS)/Rate,"")
        try:
            rate_num = float(rate) if rate else 0
            revenue  = (igst + cgst + sgst + cess) / rate_num if rate_num else ''
        except:
            revenue = ''

        ws.cell(r, 1, hsn_code).font = data_font()
        ws.cell(r, 2, hsn_desc).font = data_font()

        rate_cell = ws.cell(r, 3, rate)
        rate_cell.font = data_font()
        rate_cell.number_format = FMT_PCT

        for c_idx, val in [(4, igst), (5, cgst), (6, sgst), (7, cess)]:
            cell = ws.cell(r, c_idx, val if val else None)
            cell.font = data_font()
            cell.number_format = FMT_INT

        rev_cell = ws.cell(r, 8, revenue if revenue != '' else None)
        rev_cell.font = data_font()
        rev_cell.number_format = FMT_INT

        r += 1

    LAST_DATA = r - 1
    TOTAL_ROW = r

    # Total row — use actual SUM values (standalone file)
    ws.cell(TOTAL_ROW, 1, 'TOTAL').font = bold_font()
    
    col_totals = {4: 0, 5: 0, 6: 0, 7: 0, 8: 0}
    for row_num in range(DATA_START, TOTAL_ROW):
        for c_idx in [4, 5, 6, 7, 8]:
            v = ws.cell(row_num, c_idx).value
            try: col_totals[c_idx] += float(v or 0)
            except: pass
    
    for c_idx in [4, 5, 6, 7, 8]:
        cell = ws.cell(TOTAL_ROW, c_idx, col_totals[c_idx] if col_totals[c_idx] else None)
        cell.font = bold_font()
        cell.number_format = FMT_INT
    apply_total_row(ws, TOTAL_ROW, 1, 8)

    set_heights(ws, TOTAL_ROW)
    ws.sheet_view.showGridLines = False
    return wb, LAST_DATA, TOTAL_ROW

# ══════════════════════════════════════════════════════════
# STEP 5 — D110 VS GSTR-1
# From VBA:
# B4 = HSN Summary!H(lastRow)  → Revenue
# C4 = HSN Summary!D(lastRow)  → IGST
# D4 = HSN Summary!E(lastRow)  → CGST
# E4 = HSN Summary!F(lastRow)  → SGST
# F4 = HSN Summary!G(lastRow)  → CESS
# B5-F5 = SUMIFS on Proc_D110
# B9  = NON GST SUMIFS
# B10 = OTHERS SUMIFS
# B12 = SUM(B9:B10)
# B14 = B12 + SUM(B5:F5)
# B16 = SUM(B4:F4)
# B18 = B14 - B16
# B21 = SUM(B18:B20)
# ══════════════════════════════════════════════════════════
def make_d110_vs_gstr1(processed, new_hdrs):
    ci = {str(h).strip(): i for i, h in enumerate(new_hdrs) if h}

    # Calculate all values from processed data
    totals = dict(rev=0, igst=0, cgst=0, sgst=0, cess=0, non_gst=0, others=0)
    hsn_pivot = defaultdict(lambda: defaultdict(float))

    for row in processed:
        th = str(row[ci.get('Tax Head', 0)] or '').strip().upper()

        try: debit = float(row[ci.get('FT_DEBIT', 11)] or 0)
        except: debit = 0

        rate     = row[ci.get('Rate', 0)]
        hsn_code = str(row[ci.get('HSN Code', 0)] or '').strip()
        hsn_desc = str(row[ci.get('HSN Description', 0)] or '').strip()

        if th == 'REVENUE':   totals['rev']     += debit
        elif th == 'IGST':    totals['igst']    += debit
        elif th == 'CGST':    totals['cgst']    += debit
        elif th == 'SGST':    totals['sgst']    += debit
        elif th == 'CESS':    totals['cess']    += debit
        elif th == 'NON GST': totals['non_gst'] += debit
        elif th == 'OTHERS':  totals['others']  += debit

        if th in ['IGST','CGST','SGST','CESS']:
            hsn_pivot[(hsn_code, hsn_desc, rate)][th] += debit

    # HSN Summary revenue total (back-calc, same as make_hsn_summary)
    hsn_rev_total = 0
    for (hsnc, hsnd, rate), tv in hsn_pivot.items():
        try:
            rn = float(rate) if rate else 0
            if rn:
                hsn_rev_total += (tv.get('IGST',0) + tv.get('CGST',0) +
                                  tv.get('SGST',0) + tv.get('CESS',0)) / rn
        except:
            pass

    gstr1_rev  = hsn_rev_total
    gstr1_igst = totals['igst']
    gstr1_cgst = totals['cgst']
    gstr1_sgst = totals['sgst']
    gstr1_cess = totals['cess']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'D110_Vs_GSTR_1'

    for col, w in [('A',21.57),('B',14.71),('C',10.29),('D',13.71),('E',13.71),('F',7.29)]:
        ws.column_dimensions[col].width = w

    ws.sheet_view.showGridLines = False

    # Row 3 headers bold
    for c, h in enumerate(['Particular','Revenue','IGST','CGST','SGST','CESS'], 1):
        ws.cell(3, c, h).font = bold_font()

    def fmt_cell(r, c, val, is_bold=False):
        cell = ws.cell(r, c, val)
        cell.font = bold_font() if is_bold else data_font()
        cell.number_format = FMT_COMMA

    # Row 4 — As per GSTR-1
    ws.cell(4, 1, 'As per GSTR-1').font = data_font()
    fmt_cell(4, 2, gstr1_rev)
    fmt_cell(4, 3, gstr1_igst)
    fmt_cell(4, 4, gstr1_cgst)
    fmt_cell(4, 5, gstr1_sgst)
    fmt_cell(4, 6, gstr1_cess)

    # Row 5 — As Per D110 (SUMIFS from Proc_D110)
    ws.cell(5, 1, 'As Per D110').font = data_font()
    fmt_cell(5, 2, totals['rev'])
    fmt_cell(5, 3, totals['igst'])
    fmt_cell(5, 4, totals['cgst'])
    fmt_cell(5, 5, totals['sgst'])
    fmt_cell(5, 6, totals['cess'])

    # Row 6 — Differences (formula)
    ws.cell(6, 1, 'Differences').font = bold_font()
    for c in range(2, 7):
        cell = ws.cell(6, c, f'={get_column_letter(c)}4-{get_column_letter(c)}5')
        cell.font = bold_font()
        cell.number_format = FMT_COMMA

    # Row 9 — NON GST
    ws.cell(9,  1, 'NON GST').font = data_font()
    fmt_cell(9, 2, totals['non_gst'])

    # Row 10 — OTHERS
    ws.cell(10, 1, 'OTHERS').font = data_font()
    fmt_cell(10, 2, totals['others'])

    # Row 12 — TOTAL NO GST SUPPLIES = SUM(B9:B10)
    ws.cell(12, 1, 'TOTAL NO GST SUPPLIES').font = data_font()
    cell12 = ws.cell(12, 2, '=SUM(B9:B10)')
    cell12.font = data_font(); cell12.number_format = FMT_COMMA

    # Row 14 — NET D110 VALUE = B12 + SUM(B5:F5)
    ws.cell(14, 1, 'NET D110 VALUE').font = bold_font()
    cell14 = ws.cell(14, 2, '=B12+SUM(B5:F5)')
    cell14.font = bold_font(); cell14.number_format = FMT_COMMA

    # Row 16 — NET VALUE AS PER GSTR-1 = SUM(B4:F4)
    ws.cell(16, 1, 'NET VALUE AS PER GSTR-1').font = bold_font()
    cell16 = ws.cell(16, 2, '=SUM(B4:F4)')
    cell16.font = bold_font(); cell16.number_format = FMT_COMMA

    # Row 18 — DIFFERENCES = B14 - B16
    ws.cell(18, 1, 'DIFFERENCES').font = data_font()
    ws.cell(18, 2, '=B14-B16').number_format = FMT_COMMA

    # Row 19 — NO GST SUPPLIES = -B12
    ws.cell(19, 1, 'NO GST SUPPLIES').font = data_font()
    ws.cell(19, 2, '=-B12').number_format = FMT_COMMA

    # Row 20 — INCORRECT TAX CHARGED = B6 (Revenue difference)
    ws.cell(20, 1, 'INCORRECT TAX CHARGED').font = data_font()
    ws.cell(20, 2, '=B6').number_format = FMT_COMMA

    # Row 21 — NET DIFFERENCE = SUM(B18:B20) — yellow bg, blue font
    ws.cell(21, 1, 'NET DIFFERENCE').font = bold_font()
    cell21 = ws.cell(21, 2, '=SUM(B18:B20)')
    cell21.font = Font(name='Calibri', size=10, bold=True, color='FF0000FF')
    cell21.fill = PatternFill('solid', fgColor='FFFFFF00')
    cell21.number_format = FMT_COMMA

    set_heights(ws, 21)
    return wb

# ══════════════════════════════════════════════════════════
# SHARED: load all files and process D110
# ══════════════════════════════════════════════════════════
def load_and_process():
    hid = get_hotel_id()
    d110_bytes = get_file_bytes(hid, 'd110', 'd110')
    back_bytes = get_file_bytes(hid, 'backend', 'backend')
    einv_bytes = get_file_bytes(hid, 'einv', 'einv')
    if not d110_bytes or not back_bytes:
        raise ValueError('D110 + Backend required')
    d110_rows    = get_sheet(read_wb(d110_bytes), 'D110 Dr.Base')
    backend_rows = get_sheet(read_wb(back_bytes), 'Backend Trx,SAC_Mapping')
    gstr_rows    = get_sheet(read_wb(einv_bytes), 'GSTR-4A_EINV') if einv_bytes else []
    mapping  = build_map(backend_rows)
    einv_map = build_einv_map(gstr_rows)
    processed, new_hdrs = process_d110(d110_rows, mapping, einv_map)
    return processed, new_hdrs, einv_map

def send_wb(wb, filename):
    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return send_file(out, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ══════════════════════════════════════════════════════════
# ROUTES
# ══════════════════════════════════════════════════════════
@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'GET':
        if 'user_id' in session: return redirect('/')
        return render_template('login.html')
    data = request.get_json() or request.form
    user = get_db().execute("SELECT * FROM users WHERE username=? AND password_hash=?",
                            (data.get('username',''), hash_pw(data.get('password','')))).fetchone()
    if not user: return jsonify({'error':'Invalid username or password'}), 401
    session['user_id'] = user['id']; session['username'] = user['username']
    session['full_name'] = user['full_name']; session['role'] = user['role']
    return jsonify({'success':True, 'redirect':'/'})

@app.route('/logout')
def logout():
    session.clear(); return redirect('/login')

@app.route('/api/me')
@login_required
def me():
    conn = get_db()
    hotels = [dict(h) for h in conn.execute("SELECT * FROM hotels").fetchall()]
    conn.close()
    return jsonify({'user_id':session['user_id'],'username':session['username'],
                    'full_name':session['full_name'],'role':session['role'],'hotels':hotels})

@app.route('/api/files/<int:hotel_id>')
@login_required
def list_files(hotel_id):
    conn = get_db()
    rows = conn.execute("SELECT id,file_type,filename,file_size,uploaded_at FROM saved_files WHERE hotel_id=?",
                        (hotel_id,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/files/upload', methods=['POST'])
@login_required
def upload_file():
    hotel_id = int(request.form.get('hotel_id', 1))
    file_type = request.form.get('file_type')
    f = request.files.get('file')
    if not f or not file_type: return jsonify({'error':'Missing file or type'}), 400
    data = f.read()
    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO saved_files (hotel_id,file_type,filename,file_data,file_size,uploaded_by) VALUES (?,?,?,?,?,?)",
                 (hotel_id, file_type, f.filename, data, len(data), session.get('user_id',1)))
    conn.commit(); conn.close()
    return jsonify({'success':True, 'filename':f.filename, 'size':len(data)})

@app.route('/api/files/delete/<int:hotel_id>/<file_type>', methods=['DELETE'])
@login_required
def delete_file(hotel_id, file_type):
    conn = get_db()
    conn.execute("DELETE FROM saved_files WHERE hotel_id=? AND file_type=?", (hotel_id, file_type))
    conn.commit(); conn.close()
    return jsonify({'success':True})

@app.route('/api/history')
@login_required
def get_history():
    conn = get_db()
    rows = conn.execute("SELECT * FROM history ORDER BY processed_at DESC LIMIT 50").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/admin/users', methods=['GET'])
@login_required
def get_users():
    conn = get_db()
    users = [dict(u) for u in conn.execute("SELECT id,username,full_name,role,created_at FROM users").fetchall()]
    conn.close()
    return jsonify(users)

@app.route('/api/admin/users', methods=['POST'])
@login_required
def create_user():
    data = request.get_json()
    try:
        conn = get_db()
        conn.execute("INSERT INTO users (username,password_hash,full_name,role) VALUES (?,?,?,?)",
                     (data['username'], hash_pw(data['password']), data.get('full_name',''), data.get('role','user')))
        conn.commit(); conn.close()
        return jsonify({'success':True})
    except sqlite3.IntegrityError:
        return jsonify({'error':'Username already exists'}), 400

@app.route('/api/admin/users/<int:uid>', methods=['DELETE'])
@login_required
def delete_user(uid):
    conn = get_db()
    conn.execute("DELETE FROM users WHERE id=?", (uid,))
    conn.commit(); conn.close()
    return jsonify({'success':True})

@app.route('/api/admin/users/<int:uid>/password', methods=['POST'])
@login_required
def change_pw(uid):
    pw = request.get_json().get('password','')
    conn = get_db()
    conn.execute("UPDATE users SET password_hash=? WHERE id=?", (hash_pw(pw), uid))
    conn.commit(); conn.close()
    return jsonify({'success':True})

@app.route('/api/admin/hotels', methods=['GET'])
@login_required
def get_hotels():
    conn = get_db()
    hotels = [dict(h) for h in conn.execute("SELECT * FROM hotels").fetchall()]
    conn.close()
    return jsonify(hotels)

@app.route('/api/admin/hotels', methods=['POST'])
@login_required
def create_hotel():
    data = request.get_json()
    try:
        conn = get_db()
        conn.execute("INSERT INTO hotels (name,code,gstin) VALUES (?,?,?)",
                     (data['name'], data['code'].upper(), data.get('gstin','')))
        conn.commit(); conn.close()
        return jsonify({'success':True})
    except sqlite3.IntegrityError:
        return jsonify({'error':'Hotel code exists'}), 400

@app.route('/api/admin/hotels/<int:hid>', methods=['DELETE'])
@login_required
def delete_hotel(hid):
    conn = get_db()
    conn.execute("DELETE FROM hotels WHERE id=?", (hid,))
    conn.commit(); conn.close()
    return jsonify({'success':True})

@app.route('/admin')
@login_required
def admin_panel():
    return render_template('admin.html')

@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/process/step1', methods=['POST'])
@login_required
def step1():
    try:
        processed, new_hdrs, einv_map = load_and_process()
        wb, _, _ = make_proc_d110(None, None, None)
        # Re-build from processed
        wbo = openpyxl.Workbook()
        ws = wbo.active; ws.title = 'Proc_D110'
        for c,h in enumerate(new_hdrs,1):
            cell=ws.cell(1,c,h); cell.font=hdr_font(); cell.fill=hdr_fill()
        for r,row in enumerate(processed,2):
            for c,v in enumerate(row,1): ws.cell(r,c,v).font=data_font()
        set_heights(ws,len(processed)+1); ws.sheet_view.showGridLines=False
        save_history('D110','Step 1 - Proc D110','Proc_D110.xlsx')
        return send_wb(wbo, 'Proc_D110.xlsx')
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/process/step2', methods=['POST'])
@login_required
def step2():
    try:
        processed, new_hdrs, einv_map = load_and_process()
        wb = make_sales_register(processed, new_hdrs, einv_map)
        return send_wb(wb, 'Sales_Register.xlsx')
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/process/step3', methods=['POST'])
@login_required
def step3():
    try:
        processed, new_hdrs, _ = load_and_process()
        wb = make_sr_no_gst(processed, new_hdrs)
        return send_wb(wb, 'SR_No_GST.xlsx')
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/process/step4', methods=['POST'])
@login_required
def step4():
    try:
        processed, new_hdrs, _ = load_and_process()
        wb, _, _ = make_hsn_summary(processed, new_hdrs)
        return send_wb(wb, 'HSN_Summary.xlsx')
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/process/step5', methods=['POST'])
@login_required
def step5():
    try:
        processed, new_hdrs, _ = load_and_process()
        wb = make_d110_vs_gstr1(processed, new_hdrs)
        return send_wb(wb, 'D110_Vs_GSTR1_Reco.xlsx')
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ══════════════════════════════════════════════════════════
# D140 — JOURNAL BY CASHIER WORKING
# ══════════════════════════════════════════════════════════

def build_d140_mapping(backend_rows):
    """
    Backend: Row0=blank/date, Row1=headers, Data from Row2
    Col0=TRX Code, Col1=TRX_DESC, Col2=Tax Head, Col3=Rate, Col4=HSN Desc, Col5=HSN Code, Col6=BOF Code
    Key = TRX_DESC (lookup by description like VBA XLOOKUP on col B)
    """
    mapping = {}
    for row in backend_rows[2:]:
        if not row[1]: continue
        key = str(row[1] or '').strip().lower()
        mapping[key] = {
            'tax_head': row[2],
            'rate':     row[3],
            'hsn_desc': row[4],
            'hsn_code': row[5],
            'bof_code': row[6],
        }
    return mapping


def process_d140(d140_rows, mapping):
    """
    VBA logic:
    1. Copy D140 Dr cols A:AH (first 34 cols, delete AI:AT which are blank trailing cols)
    2. Add 8 new cols: Tax Head, Rate, HSN Desc, HSN Code, BOF Code, Rate Check, Status, Guest+Room
    
    Lookup key = TRX_DESC (col I = index 8)
    
    Rate Check (AN):
      - rate = "NO GST" → 0
      - rate = 0 or blank → PRINT_CASHIER_DEBIT(col AH=index33) * -1
      - else → PRINT_CASHIER_DEBIT / rate
    
    Guest+Room (AP) = GUEST_FULL_NAME(col G=index6) + "-" + ROOM(col R=index17)
    
    Status (AO) = SUMIF on Guest+Room key → Rate Check sum
      if abs(sum) <= 50 → "Passed" else "Error"
    """
    # Keep only first 34 cols (A:AH), discard AI:AT (blank trailing cols)
    KEEP_COLS = 34

    headers = list(d140_rows[0])[:KEEP_COLS] + [
        'Tax Head', 'Rate', 'HSN Description', 'HSN Code', 'BOF Code',
        'Rate Check', 'Status', 'Guest+Room'
    ]

    processed = []

    # Index constants (0-based)
    I_TRX_DESC   = 8   # col I
    I_GUEST      = 6   # col G = GUEST_FULL_NAME
    I_ROOM       = 17  # col R = ROOM
    I_PRINT_DR   = 33  # col AH = PRINT_CASHIER_DEBIT (last kept col)

    for row in d140_rows[1:]:
        if not any(row): continue

        base = list(row)[:KEEP_COLS]
        # pad if row shorter than 34
        while len(base) < KEEP_COLS:
            base.append(None)

        # Lookup by TRX_DESC
        trx_desc = str(base[I_TRX_DESC] or '').strip().lower()
        m = mapping.get(trx_desc, {})

        tax_head = m.get('tax_head', None)
        rate     = m.get('rate', None)
        hsn_desc = m.get('hsn_desc', None)
        hsn_code = m.get('hsn_code', None)
        bof_code = m.get('bof_code', None)

        # Rate Check — col AH = PRINT_CASHIER_DEBIT
        try:
            ah_val = float(base[I_PRINT_DR] or 0)
        except:
            ah_val = 0

        rate_str = str(rate).strip().upper() if rate is not None else ''
        if rate_str == 'NO GST':
            rate_check = 0
        elif rate_str in ('', '0', 'NONE') or rate is None or rate == 0:
            rate_check = ah_val * -1
        else:
            try:
                rate_num = float(rate)
                rate_check = ah_val / rate_num if rate_num != 0 else ah_val * -1
            except:
                rate_check = ah_val * -1

        # Guest+Room
        guest = str(base[I_GUEST] or '').strip()
        room  = str(base[I_ROOM]  or '').strip()
        guest_room = f"{guest}-{room}"

        processed.append(base + [tax_head, rate, hsn_desc, hsn_code, bof_code,
                                  rate_check, None, guest_room])  # Status filled below

    # Status — Dictionary SUMIF on Guest+Room → Rate Check sum
    from collections import defaultdict
    gr_idx   = len(headers) - 1   # Guest+Room = last col
    rc_idx   = len(headers) - 3   # Rate Check
    st_idx   = len(headers) - 2   # Status

    dict_sum = defaultdict(float)
    for row in processed:
        key = str(row[gr_idx] or '')
        try:
            dict_sum[key] += float(row[rc_idx] or 0)
        except:
            pass

    for row in processed:
        key = str(row[gr_idx] or '')
        sumval = dict_sum.get(key, 0)
        row[st_idx] = 'Passed' if abs(sumval) <= 50 else 'Error'

    return processed, headers


# ── STEP 6: PROC D140 ────────────────────────────────────
def make_proc_d140(d140_rows, backend_rows):
    mapping   = build_d140_mapping(backend_rows)
    processed, headers = process_d140(d140_rows, mapping)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Proc_D140'

    # Header row — black bg, white bold
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = hdr_font()
        cell.fill = hdr_fill()

    # Data rows
    for r, row in enumerate(processed, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v)
            cell.font = data_font()
            # Rate Check col (index 39 = col 40) — comma format
            if c == 40:
                cell.number_format = FMT_COMMA

    set_heights(ws, len(processed) + 1)
    ws.sheet_view.showGridLines = False
    return wb, processed, headers


# ── STEP 7: GST PIVOT
# ── STEP 7: GST PIVOT ─────────────────────────────────────
# VBA Macro2: Rows=HSN Code+HSN Desc+Rate, Cols=IGST/CGST/SGST/CESS
# Values=SUM(PRINT_CASHIER_DEBIT), Revenue=(IGST+CGST+SGST)/Rate
# Filter: only IGST/CGST/SGST/CESS rows (hide REVENUE/NON GST/OTHERS)
# Sort: HSN Code ASC, HSN Description ASC, Rate ASC
def make_gst_pivot(processed, headers):
    from collections import defaultdict

    idx = {h: i for i, h in enumerate(headers)}
    i_hsn_code = idx.get('HSN Code', 37)
    i_hsn_desc = idx.get('HSN Description', 36)
    i_rate     = idx.get('Rate', 35)
    i_th       = idx.get('Tax Head', 34)
    i_print_dr = idx.get('PRINT_CASHIER_DEBIT', 33)

    TAX_SHOW = {'IGST', 'CGST', 'SGST', 'CESS'}
    TAX_COLS = ['IGST', 'CGST', 'SGST', 'CESS']

    # Only include rows with IGST/CGST/SGST/CESS tax head
    pivot   = defaultdict(lambda: {t: 0.0 for t in TAX_COLS})
    key_set = set()

    for row in processed:
        th = str(row[i_th] or '').strip().upper()
        if th not in TAX_SHOW:
            continue
        hsn_code = row[i_hsn_code]
        hsn_desc = str(row[i_hsn_desc] or '').strip()
        rate     = row[i_rate]
        try:    val = float(row[i_print_dr] or 0)
        except: val = 0
        key = (hsn_code, hsn_desc, rate)
        key_set.add(key)
        pivot[key][th] += val

    # Sort: HSN Code ASC → HSN Desc ASC → Rate ASC
    key_order = sorted(key_set,
                       key=lambda k: (int(k[0]) if k[0] else 0,
                                      str(k[1]),
                                      float(k[2] or 0)))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'GST_Pivot'

    # Row 1: title
    ws.cell(1, 1, 'Sum of PRINT_CASHIER_DEBIT').font = data_font()
    ws.cell(1, 4, 'Tax Head').font = data_font()

    # Row 2: headers — black bg, white bold
    hdr_vals = ['HSN Code', 'HSN Description', 'Rate', 'IGST', 'CGST', 'SGST', 'CESS', 'Revenue']
    for c, h in enumerate(hdr_vals, 1):
        cell = ws.cell(2, c, h)
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')
        cell.fill = hdr_fill()

    # Data rows from row 3
    DATA_START = 3
    for r, key in enumerate(key_order, DATA_START):
        hsn_code, hsn_desc, rate = key
        vals = pivot[key]
        igst = vals['IGST'] if vals['IGST'] != 0 else None
        cgst = vals['CGST'] if vals['CGST'] != 0 else None
        sgst = vals['SGST'] if vals['SGST'] != 0 else None
        cess = vals['CESS'] if vals['CESS'] != 0 else None
        # Revenue = (IGST+CGST+SGST) / Rate  [CESS excluded per VBA RC[-4]:RC[-2]]
        try:
            rev = (vals['IGST'] + vals['CGST'] + vals['SGST']) / float(rate) if rate else None
        except:
            rev = None

        for c, v in enumerate([hsn_code, hsn_desc, rate, igst, cgst, sgst, cess, rev], 1):
            cell = ws.cell(r, c, v)
            cell.font = data_font()
            if c >= 4:
                cell.number_format = FMT_COMMA

    # Total row
    DATA_END  = DATA_START + len(key_order) - 1
    TOTAL_ROW = DATA_END + 1
    ws.cell(TOTAL_ROW, 1, 'Total').font = Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')
    ws.cell(TOTAL_ROW, 1).fill = hdr_fill()
    for c in range(4, 9):
        cl   = get_column_letter(c)
        cell = ws.cell(TOTAL_ROW, c, f'=SUM({cl}{DATA_START}:{cl}{DATA_END})')
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')
        cell.fill = hdr_fill()
        cell.number_format = FMT_COMMA

    set_heights(ws, TOTAL_ROW)
    ws.sheet_view.showGridLines = False
    return wb


# ── STEP 8: GL SETUP / PROC GL ───────────────────────────
# VBA Macro3 exact logic:
# 1. Add 5 cols to GL: AM=Ledger Type, AN=Ledger Name, AO=LTD Balance(blank),
#    AP=Opera/Manual(blank), AQ=Remarks
# 2. Insert 5 rows at top of Proc GL sheet
# Row 1: "GL Summary - Proc GL" title
# Row 2: blank
# Row 3: Sub Total (Z and AD cols summed)
# Row 4: Total (Z and AD cols summed)
# Row 5: blank
# Row 6: Headers
# Row 7+: Data
def _get_ledger_info(acc_code, acc_desc, user_batch, trx_desc):
    """VBA Select Case + remarks logic"""
    acc_code   = str(acc_code   or '').strip()
    acc_desc   = str(acc_desc   or '').strip().upper()
    user_batch = str(user_batch or '').strip()
    trx_desc   = str(trx_desc  or '').strip().upper()

    gst_map = {
        '217576': ('GST Payable Ledger',    'Output Central GST'),
        '217577': ('GST Payable Ledger',    'Output State GST'),
        '217578': ('GST Payable Ledger',    'Output Integrated GST'),
        '217580': ('GST Payable Ledger',    'Output Cess'),
        '217586': ('GST Payable Ledger',    'Output RCM Union GST Liab'),
        '137506': ('GST Receivable Ledger', 'Input Central GST'),
        '137507': ('GST Receivable Ledger', 'Input State GST'),
        '137508': ('GST Receivable Ledger', 'Input Integrated GST'),
        '137528': ('GST Receivable Ledger', 'Input RCM - Integrated GST'),
    }

    ledger_type = ''; ledger_name = ''
    if acc_code in gst_map:
        ledger_type, ledger_name = gst_map[acc_code]
    elif acc_code.startswith('3'):
        ledger_type = 'Revenue Ledger';  ledger_name = 'Revenue Ledger'
    elif acc_code:
        ledger_type = 'Other than GST';  ledger_name = 'Other than GST'

    gst_kw = ('GST', 'TAX', 'JV', 'ENTRY', 'CORRECT', 'PAYMENT', 'ADJUST')
    remarks = ''
    if ledger_type == 'Revenue Ledger':
        if any(w in acc_desc for w in ('BEER', 'WINE', 'LIQUOR')):
            remarks = 'Revenue-Non GST'
        else:
            remarks = 'Revenue Booked'
    elif ledger_type == 'GST Payable Ledger':
        remarks = ledger_name if user_batch else 'GST Payment & Adjustment'
    elif ledger_type == 'GST Receivable Ledger':
        remarks = 'GST Payment & Adjustment' if any(w in trx_desc for w in gst_kw) else 'GST Input Credit'
    elif ledger_type == 'Other than GST':
        remarks = 'Other than GST'

    return ledger_type, ledger_name, remarks


def make_proc_gl(gl_rows):
    # Original GL has 38 cols (0-37); 5 new cols added = 43 total
    GL_ORIG_COLS = 38
    TOTAL_COLS   = 42  # AP = col 42 (AQ removed)

    # GL col indices (0-based)
    I_ACCT     = 6   # G = Account
    I_ACCT_DESC= 7   # H = Account Description
    I_USER     = 15  # P = User
    I_DESC     = 20  # U = Description (trxDesc in VBA)

    # New col positions (0-based in output row)
    I_LEDGER_TYPE = 38  # AM
    I_LEDGER_NAME = 39  # AN
    I_LTD_BAL    = 40  # AO — blank (not set by VBA Macro3)
    I_OPERA      = 41  # AP — blank (not set by VBA Macro3)
    I_REMARKS    = 42  # AQ

    # Build enhanced data rows
    data_rows = []
    for row in gl_rows[1:]:
        if not any(row): continue
        base = list(row)[:GL_ORIG_COLS]
        while len(base) < GL_ORIG_COLS: base.append(None)

        acc_code   = str(base[I_ACCT]      or '').strip()
        acc_desc   = str(base[I_ACCT_DESC] or '').strip()
        user_batch = str(base[I_USER]      or '').strip()
        trx_desc   = str(base[I_DESC]      or '').strip()

        lt, ln, rm = _get_ledger_info(acc_code, acc_desc, user_batch, trx_desc)
        data_rows.append(base + [lt, ln, rm, None, None])  # AO=LTD Balance(calc), AP=Opera/Manual(blank), AQ=Remarks(blank)

    # ── Build workbook ────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Proc GL'

    DATA_START = 7  # row 7 = first data row
    DATA_END   = DATA_START + len(data_rows) - 1

    # ── Row 1: Title ──────────────────────────────────────
    c1 = ws.cell(1, 1, 'GL Summary - Proc GL')
    c1.font = Font(name='Calibri', size=12, bold=True)

    # ── Row 3: Sub Total ──────────────────────────────────
    ws.cell(3, 1, 'Sub Total').font = Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')
    ws.cell(3, 1).fill = hdr_fill()
    # Z = col 26, AD = col 30
    for col in [26, 30]:
        cl = get_column_letter(col)
        c  = ws.cell(3, col, f'=SUBTOTAL(9,{cl}{DATA_START}:{cl}{DATA_END})')
        c.font = Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')
        c.fill = hdr_fill()
        c.number_format = FMT_COMMA

    # ── Row 4: Total ──────────────────────────────────────
    ws.cell(4, 1, 'Total').font = Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')
    ws.cell(4, 1).fill = hdr_fill()
    for col in [26, 30]:
        cl = get_column_letter(col)
        c  = ws.cell(4, col, f'=SUM({cl}{DATA_START}:{cl}{DATA_END})')
        c.font = Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')
        c.fill = hdr_fill()
        c.number_format = FMT_COMMA

    # ── Row 6: Headers ────────────────────────────────────
    orig_hdrs = list(gl_rows[0])[:GL_ORIG_COLS]
    while len(orig_hdrs) < GL_ORIG_COLS: orig_hdrs.append(None)
    all_hdrs  = orig_hdrs + ['Ledger Type', 'Ledger Name', 'LTD Balance', 'Opera/Manual', 'Remarks']
    for c, h in enumerate(all_hdrs, 1):
        cell = ws.cell(6, c, h)
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')
        cell.fill = hdr_fill()

    # ── Data rows from row 7 ──────────────────────────────
    for r, row in enumerate(data_rows, DATA_START):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v).font = data_font()

    # ── Row heights + no gridlines ────────────────────────
    set_heights(ws, DATA_END)
    ws.sheet_view.showGridLines = False
    return wb, gl_rows


# ── STEP 9: GL SUMMARY ───────────────────────────────────
# VBA Macro4: Pivot from Proc GL
# Row fields: Ledger Type > Ledger Name > Remarks > LTD Balance
# Col field:  Opera/Manual
# Value:      SUM(Base Amount)
# ColumnGrand=True, RowGrand=True → Grand Total row + col
# Row 3 = "Sum of Base Amount" + "Opera/Manual"
# Row 4 = pivot headers (black bg)
# Data from row 5
def make_gl_summary(gl_rows):
    from collections import defaultdict

    # Proc GL data starts from row 7 (after 5 inserted rows + header)
    # But we receive raw GL rows here — apply same _get_ledger_info logic
    # Cols from GL: 6=Account, 7=AcctDesc, 14=Source(User), 15=User(batch), 20=Description, 29=BaseAmount

    # Build pivot: (ledger_type, ledger_name, remarks, ltd_bal) → {opera_manual: sum}
    pivot    = defaultdict(lambda: defaultdict(float))
    key_order = []
    key_set   = set()

    for row in gl_rows[1:]:
        if not any(row): continue
        acc_code   = str(row[6]  or '').strip()
        acc_desc   = str(row[7]  or '').strip()
        user_batch = str(row[15] or '').strip()
        trx_desc   = str(row[20] or '').strip()
        source     = str(row[14] or '').strip()
        try:    amt = float(row[29] or 0)
        except: amt = 0

        lt, ln, rm = _get_ledger_info(acc_code, acc_desc, user_batch, trx_desc)
        if not lt: continue

        ltd_bal    = None   # LTD Balance always blank (not populated in Macro3)
        om         = source if source else None  # Opera/Manual — keep source value or blank

        key = (lt, ln, rm, ltd_bal)
        key_set.add(key)
        pivot[key][om or '(blank)'] += amt

    # Sort: GST Payable → GST Receivable → Other than GST → Revenue Ledger → by name → by remarks
    lt_order = {'GST Payable Ledger':0,'GST Receivable Ledger':1,'Other than GST':2,'Revenue Ledger':3}
    key_order = sorted(key_set, key=lambda k: (lt_order.get(k[0],9), k[1], k[2]))

    # All Opera/Manual values seen
    om_cols = sorted(set(k for v in pivot.values() for k in v.keys()))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'GL Summary'

    # ── Row 3: Title / col field header ───────────────────
    ws.cell(3, 1, 'Sum of Base Amount').font = data_font()
    ws.cell(3, 5, 'Opera/Manual').font = data_font()

    # ── Row 4: Headers — black bg, white bold ─────────────
    hdr_vals = ['Ledger Type', 'Ledger Name', 'Remarks', 'LTD Balance'] + om_cols + ['Grand Total']
    for c, h in enumerate(hdr_vals, 1):
        cell = ws.cell(4, c, h)
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')
        cell.fill = hdr_fill()

    # ── Data rows from row 5 ──────────────────────────────
    DATA_START = 5
    for r, key in enumerate(key_order, DATA_START):
        lt, ln, rm, ltd = key
        ws.cell(r, 1, lt).font = data_font()
        ws.cell(r, 2, ln).font = data_font()
        ws.cell(r, 3, rm).font = data_font()
        ws.cell(r, 4, ltd).font = data_font()  # blank

        row_total = 0
        for c, om in enumerate(om_cols, 5):
            val = pivot[key].get(om, 0) or None
            cell = ws.cell(r, c, val)
            cell.font = data_font()
            cell.number_format = FMT_COMMA
            row_total += pivot[key].get(om, 0)

        # Grand Total col
        gt_cell = ws.cell(r, 5 + len(om_cols), row_total or None)
        gt_cell.font = data_font()
        gt_cell.number_format = FMT_COMMA

    # ── Grand Total row ───────────────────────────────────
    DATA_END  = DATA_START + len(key_order) - 1
    TOTAL_ROW = DATA_END + 1
    ws.cell(TOTAL_ROW, 1, 'Grand Total').font = Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')
    ws.cell(TOTAL_ROW, 1).fill = hdr_fill()
    for c in range(5, 6 + len(om_cols)):
        cl   = get_column_letter(c)
        cell = ws.cell(TOTAL_ROW, c, f'=SUM({cl}{DATA_START}:{cl}{DATA_END})')
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')
        cell.fill = hdr_fill()
        cell.number_format = FMT_COMMA

    set_heights(ws, TOTAL_ROW)
    ws.sheet_view.showGridLines = False
    return wb


# ── STEP 10: OPERA VS GL RECO ─────────────────────────────
# Compare D140 cashier totals (Opera) vs GL SWN source totals
# Date-wise: BUSINESS_DATE × source comparison
def make_opera_vs_gl(processed, headers, gl_rows):
    from collections import defaultdict

    idx = {h: i for i, h in enumerate(headers)}
    i_date = idx.get('BUSINESS_DATE', 12)
    i_dr   = idx.get('CASHIER_DEBIT', 15)
    i_th   = idx['Tax Head']

    # D140 (Opera) - sum CASHIER_DEBIT by date and tax head
    opera_by_date = defaultdict(lambda: defaultdict(float))
    for row in processed:
        date = str(row[i_date] or '').strip()
        th   = str(row[i_th] or '').strip().upper()
        try:
            dr = float(row[i_dr] or 0)
        except:
            dr = 0
        if date:
            opera_by_date[date][th] += dr

    # GL SWN (Opera GL) - sum Base Amount by Journal Date
    gl_by_date = defaultdict(float)
    for row in gl_rows[1:]:
        if not any(row): continue
        source = str(row[14] or '').strip()
        if source != 'SWN': continue
        try:
            amt = float(row[29] or 0)
        except:
            amt = 0
        jdate = row[18]
        if jdate:
            d = str(jdate)[:10] if hasattr(jdate, 'strftime') else str(jdate)[:10]
            gl_by_date[d] += amt

    # Build reco sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Opera_VS_GL'

    hdrs = ['Date', 'As per D140 (Opera) - Revenue', 'As per D140 - Total Tax',
            'As per D140 - Total', 'As per GL (SWN)', 'Difference', 'Remark']
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(1, c, h)
        cell.font = hdr_font()
        cell.fill = hdr_fill()

    all_dates = sorted(set(list(opera_by_date.keys()) + list(gl_by_date.keys())))

    for r, date in enumerate(all_dates, 2):
        d140_rev = opera_by_date[date].get('REVENUE', 0)
        d140_tax = sum(v for k, v in opera_by_date[date].items() if k in ('IGST', 'CGST', 'SGST', 'CESS'))
        d140_total = sum(opera_by_date[date].values())
        gl_total   = gl_by_date.get(date, 0)
        diff       = d140_total - gl_total
        remark     = 'OK' if abs(diff) <= 50 else 'Check'

        vals = [date, d140_rev or None, d140_tax or None, d140_total or None, gl_total or None, diff or None, remark]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.font = data_font()
            if c in [2, 3, 4, 5, 6]:
                cell.number_format = FMT_COMMA
            if c == 7:
                if v == 'OK':
                    cell.font = Font(name='Calibri', size=10, color='FF0070C0')
                else:
                    cell.font = Font(name='Calibri', size=10, color='FFC00000')

    TOTAL_ROW = len(all_dates) + 2
    ws.cell(TOTAL_ROW, 1, 'TOTAL').font = bold_font()
    ws.cell(TOTAL_ROW, 1).fill = hdr_fill()
    for c in [2, 3, 4, 5, 6]:
        cl = get_column_letter(c)
        cell = ws.cell(TOTAL_ROW, c, f'=SUM({cl}2:{cl}{TOTAL_ROW-1})')
        cell.font = bold_font(); cell.fill = hdr_fill()
        cell.number_format = FMT_COMMA

    set_heights(ws, TOTAL_ROW)
    ws.sheet_view.showGridLines = False
    return wb


# ══════════════════════════════════════════════════════════
# D140 ROUTES
# ══════════════════════════════════════════════════════════

def load_d140():
    hid = get_hotel_id()
    d140_bytes = get_file_bytes(hid, 'd140', 'd140')
    back_bytes = get_file_bytes(hid, 'backend', 'backend')
    if not d140_bytes or not back_bytes:
        return None, None, None, 'D140 + Backend required'
    d140_rows    = get_sheet(read_wb(d140_bytes), 'D140 Dr')
    backend_rows = get_sheet(read_wb(back_bytes), 'Backend Trx,SAC_Mapping')
    return d140_rows, backend_rows, None, None


@app.route('/process/d140/step1', methods=['POST'])
@login_required
def d140_step1():
    try:
        d140_rows, backend_rows, _, err = load_d140()
        if err: return jsonify({'error': err}), 400
        wb, _, _ = make_proc_d140(d140_rows, backend_rows)
        return send_wb(wb, 'Proc_D140.xlsx')
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/process/d140/step2', methods=['POST'])
@login_required
def d140_step2():
    try:
        d140_rows, backend_rows, _, err = load_d140()
        if err: return jsonify({'error': err}), 400
        mapping = build_d140_mapping(backend_rows)
        processed, headers = process_d140(d140_rows, mapping)
        wb = make_gst_pivot(processed, headers)
        return send_wb(wb, 'GST_Pivot.xlsx')
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/process/d140/step3', methods=['POST'])
@login_required
def d140_step3():
    try:
        hid = get_hotel_id()
        gl_bytes = get_file_bytes(hid, 'gl', 'gl')
        if not gl_bytes: return jsonify({'error': 'GL file required'}), 400
        gl_rows = get_sheet(read_wb(gl_bytes), 'GL')
        wb, _ = make_proc_gl(gl_rows)
        return send_wb(wb, 'Proc_GL.xlsx')
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/process/d140/step4', methods=['POST'])
@login_required
def d140_step4():
    try:
        hid = get_hotel_id()
        gl_bytes = get_file_bytes(hid, 'gl', 'gl')
        if not gl_bytes: return jsonify({'error': 'GL file required'}), 400
        gl_rows = get_sheet(read_wb(gl_bytes), 'GL')
        wb = make_gl_summary(gl_rows)
        return send_wb(wb, 'GL_Summary.xlsx')
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/process/d140/step5', methods=['POST'])
@login_required
def d140_step5():
    try:
        hid = get_hotel_id()
        d140_bytes = get_file_bytes(hid, 'd140', 'd140')
        back_bytes = get_file_bytes(hid, 'backend', 'backend')
        gl_bytes   = get_file_bytes(hid, 'gl', 'gl')
        if not d140_bytes or not back_bytes or not gl_bytes:
            return jsonify({'error': 'D140 + Backend + GL required'}), 400
        d140_rows    = get_sheet(read_wb(d140_bytes), 'D140 Dr')
        backend_rows = get_sheet(read_wb(back_bytes), 'Backend Trx,SAC_Mapping')
        gl_rows      = get_sheet(read_wb(gl_bytes), 'GL')
        mapping = build_d140_mapping(backend_rows)
        processed, headers = process_d140(d140_rows, mapping)
        wb = make_opera_vs_gl(processed, headers, gl_rows)
        return send_wb(wb, 'Opera_VS_GL_Reco.xlsx')
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

init_db()

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT',5000)))
